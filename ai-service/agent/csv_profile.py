"""
Structured CSV/XLSX → short profile markdown only (not full table dump).

Original files live under ai-service/data/csv_lake/.
Profile MD is written to Documents/ai-service/ for existing ingest_secure rglob.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)

SAMPLE_ROWS = 5
TABLE_SUFFIXES = {".csv", ".xlsx"}


def _safe_stem(path: Path) -> str:
    stem = re.sub(r"[^\w\-가-힣]+", "-", path.stem, flags=re.UNICODE).strip("-")
    return stem or "table"


def csv_lake_dir(ai_root: Path) -> Path:
    return ai_root / "data" / "csv_lake"


def profile_md_dir(secure_docs_dir: Path) -> Path:
    """Short meta MD next to other converted docs (ingest picks up all *.md)."""
    return secure_docs_dir / "ai-service"


def _load_frame(path: Path):
    import polars as pl

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pl.read_csv(str(path), infer_schema_length=10_000)
    if suffix == ".xlsx":
        try:
            return pl.read_excel(str(path))
        except Exception:
            # Fallback when fastexcel is unavailable — openpyxl
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return pl.DataFrame()
            headers = [str(c) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
            data = {
                headers[i]: [r[i] if i < len(r) else None for r in rows[1:]]
                for i in range(len(headers))
            }
            return pl.DataFrame(data)
    raise ValueError(f"unsupported table: {suffix}")


def build_profile_markdown(*, path: Path, lake_rel: str) -> str:
    df = _load_frame(path)
    n_rows, n_cols = df.height, df.width
    cols = list(df.columns)
    dtypes = [str(df[c].dtype) for c in cols]
    sample = df.head(SAMPLE_ROWS)
    stem = _safe_stem(path)

    lines: list[str] = [
        "---",
        f"doc_id: csv-profile-{stem}",
        f"title: 데이터셋 안내 — {path.name}",
        "category: data_profile",
        f"source_path: {lake_rel}",
        "converted_from: profile",
        "security_level: internal",
        "---",
        "",
        f"# 데이터셋 안내: `{path.name}`",
        "",
        "이 문서는 **전체 행을 RAG에 넣지 않습니다.** 스키마·규모·샘플만 안내합니다.",
        "원본 표는 `ai-service/data/csv_lake/`에 보관됩니다.",
        "",
        f"- **행 수:** {n_rows}",
        f"- **열 수:** {n_cols}",
        f"- **원본 경로:** `{lake_rel}`",
        "",
        "## 컬럼",
        "",
        "| column | dtype |",
        "| --- | --- |",
    ]
    for c, dt in zip(cols, dtypes):
        lines.append(f"| {c} | {dt} |")
    lines.extend(["", f"## 샘플 ({min(SAMPLE_ROWS, n_rows)}행)", ""])
    if cols:
        esc = lambda x: str(x).replace("|", "\\|").replace("\n", " ")
        lines.append("| " + " | ".join(esc(c) for c in cols) + " |")
        lines.append("| " + " | ".join("---" for _ in cols) + " |")
        for row in sample.iter_rows():
            lines.append("| " + " | ".join(esc(v) for v in row) + " |")
    lines.append("")
    return "\n".join(lines)


def write_csv_profile(
    path: Path,
    *,
    secure_docs_dir: Path,
    ai_root: Path,
) -> Path | None:
    if path.suffix.lower() not in TABLE_SUFFIXES:
        return None
    if not path.is_file():
        return None
    try:
        rel = str(path.relative_to(ai_root)).replace("\\", "/")
    except ValueError:
        rel = str(path).replace("\\", "/")
    try:
        md = build_profile_markdown(path=path, lake_rel=rel)
    except Exception as exc:  # noqa: BLE001
        logger.warning("[csv_profile] extract fail %s: %s", path.name, exc)
        return None
    out_dir = profile_md_dir(secure_docs_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{_safe_stem(path)}-profile.md"
    out.write_text(md, encoding="utf-8")
    logger.info("[csv_profile] wrote %s", out)
    return out
